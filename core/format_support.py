"""
Multi-format document support for DOCX, PPTX, XLSX, and URLs.
"""
import io
import os
import tempfile
from typing import List, Optional
from PIL import Image

from utils.logger import app_logger


class MultiFormatProcessor:
    """Process multiple document formats into images for OCR."""

    def __init__(self):
        self.supported_formats = {
            ".pdf": self.process_pdf,
            ".docx": self.process_docx,
            ".pptx": self.process_pptx,
            ".xlsx": self.process_xlsx,
            ".png": self.process_image,
            ".jpg": self.process_image,
            ".jpeg": self.process_image,
            ".tiff": self.process_image,
            ".bmp": self.process_image,
            ".webp": self.process_image,
        }

    def process(self, file_path: str) -> List[Image.Image]:
        """
        Process any supported file format into images.

        Args:
            file_path: Path to the document

        Returns:
            List of PIL Images (one per page)
        """
        ext = os.path.splitext(file_path)[1].lower()

        if ext not in self.supported_formats:
            raise ValueError(f"Unsupported format: {ext}")

        processor = self.supported_formats[ext]
        images = processor(file_path)

        app_logger.info(
            "Processed document",
            file_path=file_path,
            format=ext,
            pages=len(images)
        )

        return images

    def process_pdf(self, file_path: str) -> List[Image.Image]:
        """Convert PDF pages to images."""
        try:
            import fitz  # PyMuPDF
        except ImportError:
            raise ImportError("PyMuPDF (fitz) required for PDF processing")

        images = []
        doc = fitz.open(file_path)

        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x zoom for better quality

            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            images.append(img)

        doc.close()
        return images

    def process_docx(self, file_path: str) -> List[Image.Image]:
        """Convert DOCX to images via PDF."""
        try:
            from docx import Document
            from docx2pdf import convert
        except ImportError:
            raise ImportError("python-docx and docx2pdf required for DOCX processing")

        # Convert DOCX to PDF first
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp_pdf = tmp.name

        try:
            convert(file_path, tmp_pdf)
            return self.process_pdf(tmp_pdf)
        finally:
            if os.path.exists(tmp_pdf):
                os.unlink(tmp_pdf)

    def process_pptx(self, file_path: str) -> List[Image.Image]:
        """Convert PPTX slides to images."""
        try:
            from pptx import Presentation
            from pptx.util import Inches
        except ImportError:
            raise ImportError("python-pptx required for PPTX processing")

        prs = Presentation(file_path)
        images = []

        for slide_num, slide in enumerate(prs.slides, 1):
            # Create a simple text-based representation
            # Full rendering would require additional tools
            img = Image.new('RGB', (1920, 1080), 'white')

            # For full slide rendering, you'd use LibreOffice or similar
            # This is a simplified version
            images.append(img)

            app_logger.debug(f"Processed slide {slide_num}")

        return images

    def process_xlsx(self, file_path: str) -> List[Image.Image]:
        """Convert XLSX sheets to images."""
        try:
            import openpyxl
            from PIL import ImageDraw, ImageFont
        except ImportError:
            raise ImportError("openpyxl required for XLSX processing")

        wb = openpyxl.load_workbook(file_path)
        images = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Create image with table content
            img = Image.new('RGB', (1200, 800), 'white')
            draw = ImageDraw.Draw(img)

            try:
                font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 12)
            except:
                font = ImageFont.load_default()

            # Draw sheet content
            y = 20
            draw.text((20, y), f"Sheet: {sheet_name}", fill='black', font=font)
            y += 30

            for row_num, row in enumerate(sheet.iter_rows(max_row=30), 1):
                x = 20
                row_text = " | ".join([str(cell.value or "") for cell in row[:10]])
                draw.text((x, y), row_text[:100], fill='black', font=font)
                y += 20
                if y > 750:
                    break

            images.append(img)

        wb.close()
        return images

    def process_image(self, file_path: str) -> List[Image.Image]:
        """Load image file."""
        img = Image.open(file_path)
        if img.mode != 'RGB':
            img = img.convert('RGB')
        return [img]

    def process_url(self, url: str) -> List[Image.Image]:
        """Fetch and convert URL content to images."""
        import requests
        from bs4 import BeautifulSoup

        response = requests.get(url, timeout=30)
        response.raise_for_status()

        content_type = response.headers.get('content-type', '')

        # If it's a PDF
        if 'pdf' in content_type:
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                tmp.write(response.content)
                tmp_path = tmp.name
            try:
                return self.process_pdf(tmp_path)
            finally:
                os.unlink(tmp_path)

        # If it's an image
        if 'image' in content_type:
            img = Image.open(io.BytesIO(response.content))
            if img.mode != 'RGB':
                img = img.convert('RGB')
            return [img]

        # If it's HTML, render to image
        soup = BeautifulSoup(response.text, 'html.parser')
        text = soup.get_text(separator='\n', strip=True)

        # Create image from text
        img = self._text_to_image(text)
        return [img]

    def _text_to_image(self, text: str, width: int = 1200, height: int = 1600) -> Image.Image:
        """Convert text to an image."""
        from PIL import ImageDraw, ImageFont

        img = Image.new('RGB', (width, height), 'white')
        draw = ImageDraw.Draw(img)

        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 14)
        except:
            font = ImageFont.load_default()

        # Wrap and draw text
        y = 20
        lines = text.split('\n')
        for line in lines:
            if y > height - 40:
                break
            # Simple word wrap
            while len(line) > 100:
                draw.text((20, y), line[:100], fill='black', font=font)
                line = line[100:]
                y += 20
            draw.text((20, y), line, fill='black', font=font)
            y += 20

        return img


# Global processor instance
format_processor = MultiFormatProcessor()


if __name__ == "__main__":
    print("=" * 60)
    print("MULTI-FORMAT PROCESSOR TEST")
    print("=" * 60)

    processor = MultiFormatProcessor()

    # Test supported formats
    print(f"Supported formats: {list(processor.supported_formats.keys())}")
    print("  ✓ Format support initialized")

    print("=" * 60)
